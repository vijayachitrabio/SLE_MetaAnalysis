import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Set Paths
base_dir = "/Users/vijayachitramodhukur/Library/Mobile Documents/com~apple~CloudDocs/ECLAI/GWAs_meta_analysis/AMH_MEnopause/SLE_MetaAnalysis"
supp_dir = os.path.join(base_dir, "supplementary_files")
output_file = os.path.join(supp_dir, "Master_Supplementary_Tables.xlsx")

# Create Workbook
wb = openpyxl.Workbook()
# remove default sheet
wb.remove(wb.active)

# 1. Create Title Page
ws_title = wb.create_sheet(title="Title Page")
ws_title.views.sheetView[0].showGridLines = True

# Define exact titles
toc_data = [
    ("Supplementary Table 1", "Genome-wide significant loci from the discovery meta-analysis"),
    ("Supplementary Table 2", "Heterogeneity and cohort-level statistics for prioritized variants"),
    ("Supplementary Table 3", "Consolidated LAVA prioritization results"),
    ("Supplementary Table 4", "Bayesian colocalization results"),
    ("Supplementary Table 5", "Spanish replication results for prioritized variants"),
    ("Supplementary Table 6", "Pathway enrichment results"),
    ("Supplementary Table 7", "PheWAS summary of prioritized SLE-associated variants"),
    ("Supplementary Table 8", "Therapeutic annotation of prioritized SLE-associated genes"),
    ("Supplementary Table 9", "Statistical fine-mapping (SuSiE) of prioritized SLE-associated loci")
]

# Write TOC Header
ws_title.cell(row=1, column=1, value="Table").font = Font(bold=True, color="FFFFFF")
ws_title.cell(row=1, column=2, value="Title").font = Font(bold=True, color="FFFFFF")

# Header Fill
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
for col in [1, 2]:
    cell = ws_title.cell(row=1, column=col)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Write TOC Data
thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

for idx, (tbl, title) in enumerate(toc_data, start=2):
    c1 = ws_title.cell(row=idx, column=1, value=tbl)
    c2 = ws_title.cell(row=idx, column=2, value=title)
    c1.font = Font(bold=True)
    c1.border = thin_border
    c2.border = thin_border

# Write Abbreviations
ws_title.cell(row=12, column=1, value="Abbreviations").font = Font(bold=True, size=12)

abbrev_headers = ["Abbreviation", "Definition"]
ws_title.cell(row=13, column=1, value="Abbreviation").font = Font(bold=True, color="FFFFFF")
ws_title.cell(row=13, column=2, value="Definition").font = Font(bold=True, color="FFFFFF")
for col in [1, 2]:
    ws_title.cell(row=13, column=col).fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")

abbrev_data = [
    ("LAVA", "Local Analysis of Variant Association"),
    ("PP4", "Posterior Probability of hypothesis 4 (colocalization)"),
    ("FDR", "False Discovery Rate"),
    ("LDSC", "Linkage Disequilibrium Score Regression"),
    ("PheWAS", "Phenome-Wide Association Study"),
    ("eQTL", "expression Quantitative Trait Locus"),
    ("SuSiE", "Sum of Single Effects"),
    ("PIP", "Posterior Inclusion Probability")
]

for idx, (abb, dfn) in enumerate(abbrev_data, start=14):
    c1 = ws_title.cell(row=idx, column=1, value=abb)
    c2 = ws_title.cell(row=idx, column=2, value=dfn)
    c1.font = Font(bold=True)
    c1.border = thin_border
    c2.border = thin_border

# Set Column Widths for Title Page
ws_title.column_dimensions['A'].width = 25
ws_title.column_dimensions['B'].width = 75

# 2. Write Sheets
files = [
    "Supplementary_Table_1_Master_discovery_meta-analysis_results.tsv",
    "Supplementary_Table_2_Quality-control_metrics_for_discovery_meta-analysis.tsv",
    "Supplementary_Table_3_Consolidated_LAVA_prioritization_results.tsv",
    "Supplementary_Table_4_Bayesian_colocalization_results.tsv",
    "Supplementary_Table_5_Spanish_replication_results_for_prioritized_variants.tsv",
    "Supplementary_Table_6_Pathway_enrichment_results.tsv",
    "Supplementary_Table_7_Cross-trait_genetic_correlation_PheWAS_summary.tsv",
    "Supplementary_Table_8_Drug_and_therapeutic_target_annotation_summary.tsv",
    "Supplementary_Table_9_SuSiE_statistical_fine-mapping_results.tsv"
]

for i, filename in enumerate(files, start=1):
    sheet_name = f"Supp_Table_{i}"
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    file_path = os.path.join(supp_dir, filename)
    if os.path.exists(file_path):
        # Load TSV
        df = pd.read_csv(file_path, sep="\t")
        
        # Title in Row 1
        full_title = f"Supplementary Table {i}. {toc_data[i-1][1]}"
        ws.cell(row=1, column=1, value=full_title).font = Font(bold=True, size=13, color="1F497D")
        
        # Headers in Row 3
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        # Data in Row 4 onwards
        for row_idx, row_data in enumerate(df.values, start=4):
            for col_idx, val in enumerate(row_data, start=1):
                # Handle NA cleanly
                if pd.isna(val):
                    val_to_write = ""
                else:
                    val_to_write = val
                cell = ws.cell(row=row_idx, column=col_idx, value=val_to_write)
                cell.border = thin_border
                
        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # ignore the first row title for column width calculation
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    else:
        ws.cell(row=3, column=1, value=f"File {filename} not found.")

wb.save(output_file)
print("Master Excel workbook generated cleanly via Python openpyxl!")
