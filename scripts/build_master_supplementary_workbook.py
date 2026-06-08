from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension


BASE_DIR = Path(__file__).resolve().parents[1]
RESULTS_DIR = BASE_DIR / "results"
OUTPUT_DIR = BASE_DIR / "supplementary_files"
OUTPUT_FILE = OUTPUT_DIR / "Master_Supplementary_Tables.xlsx"
MANUSCRIPT_TITLE = (
    "Systemic lupus erythematosus genome-wide association meta-analysis"
)

TABLE_META = [
    (
        1,
        "Genome-wide significant loci from the discovery meta-analysis",
        "Lead variants and locus-level annotations from the SLE discovery meta-analysis, including local genetic correlation, prioritization, novelty, colocalization, therapeutic annotation, and functional evidence fields.",
    ),
    (
        2,
        "Heterogeneity and cohort-level statistics for prioritized variants",
        "Cohort-specific association statistics and heterogeneity metrics for prioritized variants across Bentham and FinnGen discovery inputs.",
    ),
    (
        3,
        "Consolidated LAVA prioritization results",
        "Local genetic correlation and locus prioritization results used to classify candidate SLE loci according to the validated prioritization framework.",
    ),
    (
        4,
        "Bayesian colocalization results",
        "Colocalization summary for prioritized loci and candidate genes across the primary GTEx v10 immune-relevant tissues.",
    ),
    (
        5,
        "Spanish replication results for prioritized variants",
        "Independent Spanish cohort validation statistics, including effect direction concordance and replication status for prioritized discovery variants.",
    ),
    (
        6,
        "Pathway enrichment results",
        "Ranked pathway enrichment results across curated pathway sources, including gene counts and multiple-testing-adjusted significance values.",
    ),
    (
        7,
        "PheWAS summary of prioritized SLE-associated variants",
        "Cross-trait association summary for prioritized SLE variants from GWAS Catalog lookups, including reported traits, effect estimates, and study identifiers.",
    ),
    (
        8,
        "Therapeutic annotation of prioritized SLE-associated genes",
        "Drug, pathway, and clinical-development annotations for prioritized candidate genes and therapeutic targets.",
    ),
    (
        9,
        "Statistical fine-mapping of prioritized SLE-associated loci",
        "Credible-set variant summary for prioritized SLE loci, including posterior inclusion probabilities and association Z-scores.",
    ),
]

ABBREVIATIONS = [
    ("BP", "Base-pair genomic position"),
    ("CHR", "Chromosome"),
    ("CS", "Credible set"),
    ("eQTL", "Expression quantitative trait locus"),
    ("FDR", "False discovery rate"),
    ("GTEx", "Genotype-Tissue Expression project"),
    ("LAVA", "Local Analysis of Variant Association"),
    ("PheWAS", "Phenome-wide association study"),
    ("PIP", "Posterior inclusion probability"),
    ("PP4", "Posterior probability for a shared causal variant in colocalization analysis"),
    ("RSID", "Reference SNP cluster identifier"),
]


def read_tsv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle, delimiter="\t"))


def coerce_value(value: str):
    value = value.strip()
    if value in {"", '""'}:
        return None
    try:
        if value.lower().startswith("rs"):
            return value
        if value.startswith("0") and value not in {"0"} and not value.startswith("0."):
            return value
        number = float(value)
        if number.is_integer() and "e" not in value.lower() and "." not in value:
            return int(number)
        return number
    except ValueError:
        return value


def set_widths(ws, rows: list[list[str]], max_width: int = 42) -> None:
    widths = {}
    for row in rows[: min(len(rows), 250)]:
        for idx, value in enumerate(row, start=1):
            text = "" if value is None else str(value)
            widths[idx] = max(widths.get(idx, 0), len(text))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 11), max_width)


def format_numeric_columns(ws, header: list[str], last_row: int) -> None:
    scientific_terms = {"p", "p-value", "p_meta", "lava_p", "raw p-value", "adjusted p-value/fdr"}
    decimal_terms = {"beta", "lava_rg", "pp4", "pip", "z", "odds_ratio"}
    integer_terms = {"chr", "bp", "lava_loc", "gene count", "pubmed_id"}

    for col_idx, name in enumerate(header, start=1):
        normalized = name.strip().lower()
        number_format = None
        if normalized in integer_terms:
            number_format = "0"
        elif normalized in scientific_terms or normalized.endswith("p-value"):
            number_format = "0.00E+00"
        elif normalized in decimal_terms or "beta" in normalized:
            number_format = "0.000"

        if number_format:
            for row_idx in range(5, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = number_format


def style_table_sheet(
    ws,
    table_number: int,
    title: str,
    description: str,
    rows: list[list[str]],
) -> None:
    max_col = max(len(row) for row in rows) if rows else 1
    last_row = len(rows) + 3

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(max_col)}{max(4, last_row)}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "4:4"

    title_cell = ws.cell(row=1, column=1, value=f"Supplementary Table {table_number}. {title}")
    title_cell.font = Font(bold=True, size=12, color="111111")
    title_cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.row_dimensions[1].height = 24

    desc_cell = ws.cell(row=2, column=1, value=description)
    desc_cell.font = Font(size=10, color="333333", italic=True)
    desc_cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.row_dimensions[2].height = 36

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(bold=True, color="FFFFFF")
    thin_gray = Side(style="thin", color="D9D9D9")
    header_border = Border(top=Side(style="medium", color="111111"), bottom=Side(style="medium", color="111111"))
    body_border = Border(bottom=thin_gray)

    for r_idx, row in enumerate(rows, start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=coerce_value(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r_idx == 4:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = header_border
            else:
                cell.border = body_border
                if (r_idx - 4) % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor="F7F7F7")

    set_widths(ws, rows)
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        width = ws.column_dimensions[letter].width or 12
        if width > 28:
            ws.column_dimensions[letter].width = min(width, 42)
        ws.column_dimensions[letter].bestFit = False

    if rows:
        format_numeric_columns(ws, rows[0], last_row)

    ws.row_dimensions[4].height = 34


def build_workbook() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Title Page"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Supplementary Tables"
    ws["A1"].font = Font(bold=True, size=20, color="111111")
    ws["A2"] = MANUSCRIPT_TITLE
    ws["A2"].font = Font(size=12, color="333333", italic=True)
    ws["A3"] = "Master workbook for journal submission"
    ws["A3"].font = Font(size=10, color="555555")
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    ws.merge_cells("A3:E3")

    headers = ["Table", "Worksheet", "Title", "Description", "Rows"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=Side(style="medium", color="111111"), bottom=Side(style="medium", color="111111"))

    thin = Side(style="thin", color="D1D5DB")
    for row_idx, (num, title, desc) in enumerate(TABLE_META, start=6):
        sheet_name = f"Supplementary Table {num}"
        row_count = max(0, len(read_tsv(RESULTS_DIR / f"Supplementary_Table_{num}.tsv")) - 1)
        values = [f"Supplementary Table {num}", sheet_name, title, desc, row_count]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx in {1, 2}:
                cell.font = Font(bold=col_idx == 1, color="111111")
        ws.cell(row=row_idx, column=2).hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(row=row_idx, column=2).style = "Hyperlink"

    notes_start = 17
    ws.cell(row=notes_start, column=1, value="General Notes").font = Font(
        bold=True, size=13, color="111111"
    )
    notes = [
        "High-confidence loci require both LAVA rho > 0.30 and colocalization PP4 > 0.80.",
        "Spanish replication is considered supportive when P < 0.05 with concordant direction of effect.",
        "LD reference panel: 1000 Genomes Phase 3 European ancestry.",
        "Primary colocalization tissues: GTEx v10 Whole Blood and Spleen.",
    ]
    for offset, note in enumerate(notes, start=1):
        cell = ws.cell(row=notes_start + offset, column=1, value=note)
        ws.merge_cells(start_row=notes_start + offset, start_column=1, end_row=notes_start + offset, end_column=5)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=10, color="333333")

    abbrev_start = 24
    ws.cell(row=abbrev_start, column=1, value="Abbreviations").font = Font(
        bold=True, size=13, color="111827"
    )
    for col_idx, header in enumerate(["Abbreviation", "Definition"], start=1):
        cell = ws.cell(row=abbrev_start + 1, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.font = Font(bold=True, color="FFFFFF")
    for row_idx, (abbr, definition) in enumerate(ABBREVIATIONS, start=abbrev_start + 2):
        ws.cell(row=row_idx, column=1, value=abbr).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=definition)
        for col_idx in (1, 2):
            ws.cell(row=row_idx, column=col_idx).border = Border(bottom=thin)
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 54
    ws.column_dimensions["D"].width = 82
    ws.column_dimensions["E"].width = 10
    ws.freeze_panes = "A6"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    for num, title, desc in TABLE_META:
        source = RESULTS_DIR / f"Supplementary_Table_{num}.tsv"
        table_rows = read_tsv(source)
        table_ws = wb.create_sheet(f"Supplementary Table {num}")
        style_table_sheet(table_ws, num, title, desc, table_rows)

    wb.save(OUTPUT_FILE)

    check = load_workbook(OUTPUT_FILE, read_only=False, data_only=False)
    expected_sheets = ["Title Page"] + [f"Supplementary Table {i}" for i in range(1, 10)]
    missing = [sheet for sheet in expected_sheets if sheet not in check.sheetnames]
    if missing:
        raise RuntimeError(f"Missing sheets after save: {missing}")
    for i in range(1, 10):
        ws_check = check[f"Supplementary Table {i}"]
        if ws_check.max_row < 1 or ws_check.max_column < 1:
            raise RuntimeError(f"Supplementary Table {i} appears empty")


if __name__ == "__main__":
    build_workbook()
    print(OUTPUT_FILE)
